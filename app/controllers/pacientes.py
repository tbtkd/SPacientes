from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from werkzeug.utils import secure_filename
import os
import openpyxl
from datetime import datetime
from app.models.paciente import Paciente
from app.models.pago import Pago
from app.models.historial_clinico import HistorialClinico
from app.models.valoracion_antropometrica import ValoracionAntropometrica
import sqlite3
from app.db import query_db

pacientes = Blueprint('pacientes', __name__, url_prefix='/pacientes')

@pacientes.route('/nuevo', methods=['GET', 'POST'])
def nuevo_paciente():
    if request.method == 'POST':
        try:
            # Obtener y validar datos del formulario
            nombre = request.form.get('nombre', '').strip()
            apellido_paterno = request.form.get('apellido_paterno', '').strip()
            apellido_materno = request.form.get('apellido_materno', '').strip()
            fecha_nacimiento = request.form.get('fecha_nacimiento', '').strip()
            telefono = request.form.get('telefono', '').strip()
            correo = request.form.get('correo', '').strip()
            ciudad = request.form.get('ciudad', '').strip()
            
            # Validaciones básicas
            if not all([nombre, apellido_paterno, apellido_materno, fecha_nacimiento, telefono, correo, ciudad]):
                flash('Todos los campos son obligatorios', 'error')
                return render_template('pacientes/nuevo_paciente.html')
            
            # Validar teléfono
            if len(telefono) != 10 or not telefono.isdigit():
                flash('El teléfono debe tener exactamente 10 dígitos numéricos', 'error')
                return render_template('pacientes/nuevo_paciente.html')
            
            Paciente.crear(
                nombre, apellido_paterno, 
                apellido_materno, fecha_nacimiento, 
                telefono,
                correo, ciudad
            )
            flash('Paciente registrado exitosamente', 'success')
            return redirect(url_for('pacientes.lista_pacientes'))
        except Exception as e:
            flash(f'Error al registrar el paciente: {str(e)}', 'error')
            return render_template('pacientes/nuevo_paciente.html')
    return render_template('pacientes/nuevo_paciente.html')

@pacientes.route('/')
def lista_pacientes():
    busqueda = request.args.get('busqueda', '')
    pacientes = Paciente.buscar(busqueda)
    return render_template('pacientes/lista_pacientes.html', pacientes=pacientes, busqueda=busqueda)

@pacientes.route('/<int:id>', methods=['GET', 'POST'])
def detalle_paciente(id):
    paciente = Paciente.obtener_por_id(id)
    if paciente is None:
        flash('Paciente no encontrado', 'error')
        return redirect(url_for('pacientes.lista_pacientes'))
    
    if request.method == 'POST':
        if 'fecha_pago' in request.form:
            fecha_pago = request.form['fecha_pago']
            try:
                Pago.registrar(id, fecha_pago)
                flash('Pago registrado exitosamente', 'success')
            except Exception as e:
                flash(f'Error al registrar el pago: {str(e)}', 'error')
        elif 'ultima_dieta' in request.form:
            ultima_dieta = request.form['ultima_dieta']
            try:
                exito, mensaje = ValoracionAntropometrica.actualizar_ultima_dieta(id, ultima_dieta)
                if exito:
                    flash(mensaje, 'success')
                else:
                    flash(mensaje, 'error')
            except Exception as e:
                flash(f'Error al actualizar la última dieta: {str(e)}', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))

    ultimo_pago = Pago.obtener_ultimo_pago(id)
    historial = HistorialClinico.obtener_por_paciente_id(id)
    ultima_valoracion = ValoracionAntropometrica.obtener_ultima_por_paciente(id)
    return render_template('pacientes/detalle_paciente.html', 
                            paciente=paciente, 
                            ultimo_pago=ultimo_pago,
                            historial=historial,
                            ultima_valoracion=ultima_valoracion)

@pacientes.route('/<int:id>/editar', methods=['GET', 'POST'])
def editar_paciente(id):
    paciente = Paciente.obtener_por_id(id)
    if paciente is None:
        flash('Paciente no encontrado', 'error')
        return redirect(url_for('pacientes.lista_pacientes'))
    
    if request.method == 'POST':
        try:
            Paciente.actualizar(
                id, request.form['nombre'], request.form['apellido_paterno'], 
                request.form['apellido_materno'], request.form['fecha_nacimiento'], 
                request.form['telefono'], request.form['correo'], 
                request.form['ciudad'], request.form['estatus']
            )
            flash('Información del paciente actualizada exitosamente', 'success')
            return redirect(url_for('pacientes.detalle_paciente', id=id))
        except sqlite3.IntegrityError:
            flash('Error: El correo electrónico ya está registrado', 'error')
        except Exception as e:
            flash(f'Error al actualizar el paciente: {str(e)}', 'error')
    return render_template('pacientes/editar_paciente.html', paciente=paciente)

@pacientes.route('/<int:id>/pago', methods=['POST'])
def registrar_pago_paciente(id):
    fecha_pago = request.form['fecha_pago']
    Pago.registrar(id, fecha_pago)
    flash('Pago registrado exitosamente', 'success')
    return redirect(url_for('pacientes.detalle_paciente', id=id))

def extraer_valor(texto, clave):
    """Extrae el número inmediatamente después de la clave en un texto."""
    try:
        partes = texto.lower().split(clave)
        if len(partes) > 1:
            for item in partes[1].split():
                try:
                    return float(item)
                except ValueError:
                    continue
    except Exception as e:
        print(f"Error extrayendo {clave}: {e}")
    return None

def convertir_fecha(fecha_valor):
    """Convierte un valor de celda en una fecha válida, asegurando formato correcto."""
    if isinstance(fecha_valor, datetime):
        return fecha_valor.date()
    elif isinstance(fecha_valor, str):
        try:
            partes = fecha_valor.split()
            fecha_limpia = partes[0] if len(partes) > 0 else fecha_valor
            return datetime.strptime(fecha_limpia, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None

def existe_registro(paciente_id, numero_cita, fecha):
    """Verifica si un registro con el mismo número de cita y fecha ya existe en la base de datos."""
    resultado = query_db('''SELECT id FROM valoracion_antropometrica WHERE paciente_id = ? AND numero_cita = ? AND fecha = ?''',
                         [paciente_id, numero_cita, fecha], one=True)
    return resultado is not None

@pacientes.route('/<int:id>/cargar-excel', methods=['POST'])
def cargar_excel(id):
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
    
    file = request.files['excel_file']
    if file.filename == '' or not file.filename.endswith(('.xls', '.xlsx')):
        return jsonify({'success': False, 'message': 'El archivo debe ser un Excel (.xls o .xlsx)'})

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        estatura = sheet['M8'].value if sheet['M8'].value else None
        row = 10
        registros_procesados = 0
        errores = []
        
        while True:
            if not sheet.cell(row=row, column=12).value:
                break
            
            try:
                fecha = convertir_fecha(sheet.cell(row=row, column=13).value)
                if not fecha:
                    errores.append(f"Error en fila {row}: Formato de fecha inválido ({sheet.cell(row=row, column=13).value})")
                    row += 1
                    continue
                
                numero_cita = sheet.cell(row=row, column=12).value
                if existe_registro(id, numero_cita, fecha):
                    errores.append(f"Error en fila {row}: Registro con número de cita {numero_cita} y fecha {fecha} ya existe.")
                    row += 1
                    continue
                
                datos = {
                    'paciente_id': id,
                    'numero_cita': numero_cita,
                    'fecha': fecha,
                    'estatura': estatura,
                    'peso': sheet.cell(row=row, column=14).value,
                    'imc': extraer_valor(str(sheet.cell(row=row, column=15).value or ''), 'imc'),
                    'grasa': extraer_valor(str(sheet.cell(row=row, column=15).value or ''), 'grasa'),
                    'cintura': sheet.cell(row=row, column=16).value,
                    'torax': sheet.cell(row=row, column=17).value,
                    'brazo': sheet.cell(row=row, column=18).value,
                    'cadera': sheet.cell(row=row, column=19).value,
                    'pierna': sheet.cell(row=row, column=20).value,
                    'pantorrilla': sheet.cell(row=row, column=21).value,
                    'tension_arterial': sheet.cell(row=row, column=24).value,
                    'frecuencia_cardiaca': sheet.cell(row=row, column=25).value,
                    'bicep': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'bc'),
                    'tricep': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'tc'),
                    'suprailiaco': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'si'),
                    'subescapular': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'se'),
                    'femoral': None,
                    'porcentaje_grasa': sheet.cell(row=row, column=23).value,
                    'ultima_dieta': None
                }
                
                exito, mensaje = ValoracionAntropometrica.crear(id, datos)
                if exito:
                    registros_procesados += 1
                else:
                    errores.append(f"Error en fila {row}: {mensaje}")
                
            except Exception as e:
                errores.append(f"Error en fila {row}: {str(e)}")
            
            row += 1
        
        resultado = {
            'success': True,
            'message': f'Se procesaron {registros_procesados} registros. Proceso finalizado correctamente.',
            'errores': errores if errores else "No se encontraron errores."
        }
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al procesar el archivo: {str(e)}'})
